"""XLSX generation — TOOLS.md section 11.2: "openpyxl is the final, locked
library for XLSX generation." Only `Workbook` (`.xlsx`), never a macro-
enabled workbook, so there is no macro-execution surface to sanitize
(TOOLS.md's "must not execute macros" is satisfied by construction).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def generate_xlsx(*, columns: list[str], rows: list[dict[str, object]], sheet_title: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_title[:31] or "Report"

    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(column) for column in columns])

    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, len(column) + 2)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
